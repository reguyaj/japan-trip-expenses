import openpyxl

wb = openpyxl.load_workbook('temp_copy.xlsx', data_only=True)
ws = wb['Japan Trip']

print('=== Finding pattern: store name row (no K-N) followed by amount row (has K-N) ===')
print('These cases show the WRONG name in the report:\n')

prev_txn = None
prev_row = None
prev_has_amounts = False

for r in range(2, 96):
    c_val = ws.cell(r, 3).value
    k = ws.cell(r, 11).value or 0
    l = ws.cell(r, 12).value or 0
    m = ws.cell(r, 13).value or 0
    n = ws.cell(r, 14).value or 0
    has_amounts = (k + l + m + n) > 0

    if c_val and has_amounts and prev_txn and not prev_has_amounts:
        print(f'Row {r}: Shows as "{c_val}" but parent is "{prev_txn}" (row {prev_row})')

    if c_val:
        prev_txn = c_val
        prev_row = r
        prev_has_amounts = has_amounts
    else:
        prev_has_amounts = False

print('\n=== Rows in report with person-name transactions ===')
for r in range(2, 96):
    c_val = ws.cell(r, 3).value
    k = ws.cell(r, 11).value or 0
    l = ws.cell(r, 12).value or 0
    m = ws.cell(r, 13).value or 0
    n = ws.cell(r, 14).value or 0
    if (k + l + m + n) > 0 and c_val:
        name = str(c_val).strip().lower()
        if name in ['jes', 'cha', 'joh', 'joyce'] or name.startswith(('jes ', 'cha ', 'jes-', 'cha-', 'jes -', 'cha -')):
            print(f'Row {r}: "{c_val}" (K={k}, L={l}, M={m}, N={n})')
