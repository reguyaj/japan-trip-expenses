import openpyxl

wb = openpyxl.load_workbook('temp_copy.xlsx', data_only=True)
ws = wb['Japan Trip']

# Simulate the new parsing logic with details collection
print('=== TRANSACTIONS WITH DETAILS (new logic) ===')
current_day = 0
day_label = ''
person_names = ['jes', 'cha', 'joyce', 'joh']

# First collect all rows
rows = []
for row in range(2, 97):
    b = ws[f'B{row}'].value
    c = ws[f'C{row}'].value
    k = ws[f'K{row}'].value or 0
    l = ws[f'L{row}'].value or 0
    m = ws[f'M{row}'].value or 0
    n = ws[f'N{row}'].value or 0
    has_amounts = (k + l + m + n) > 0
    rows.append({'row': row, 'b': b, 'c': str(c).strip() if c else '', 'k': k, 'l': l, 'm': m, 'n': n, 'has_amounts': has_amounts})

# Process
import re
for i, row in enumerate(rows):
    # Update day label if date found (but don't skip - might also have amounts)
    if row['b'] and str(row['b']) not in ['Date', 'OTHERS']:
        current_day += 1
        day_label = f'Day {current_day}'
    elif row['b'] and str(row['b']) == 'OTHERS':
        day_label = 'OTHERS'
    
    if not row['c']:
        continue
    if not row['has_amounts']:
        continue
    
    # Collect details (sub-items after this row until next amount row or day break)
    details = []
    for j in range(i + 1, len(rows)):
        sub = rows[j]
        if sub['b'] is not None:
            break  # hit next day
        if sub['has_amounts']:
            break  # hit next main transaction
        if not sub['c']:
            continue
        # Skip pure split markers like "/2", "/3", "/4"
        if re.match(r'^/\d+\s*$', sub['c']):
            continue
        details.append(sub['c'])
    
    detail_str = ', '.join(details) if details else '-'
    total = row['k'] + row['l'] + row['m'] + row['n']
    print(f'  {day_label:6s} | Row {row["row"]:2d} | {row["c"][:35]:35s} | Total: {total:10.2f} | Details: {detail_str}')

print(f'\nTotal transactions: {sum(1 for r in rows if r["has_amounts"])}')
