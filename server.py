"""
Japan Trip Expense Report Server
Dynamically reads the Excel file on each page load so changes are reflected immediately.
Run: python server.py
View: http://localhost:8080
"""

import http.server
import json
import os
import re
from datetime import datetime
from urllib.parse import urlparse

import openpyxl

EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "JAPAN EXPENSE.xlsx")
HTML_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "report.html")
PORT = 8080


def parse_excel():
    """Parse the Excel file and return structured expense data."""
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["Japan Trip"]

    # Load formulas to determine CC/DC rows
    wb_formulas = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
    ws_formulas = wb_formulas["Japan Trip"]
    cc_rows = set()
    dc_rows = set()
    cc_formula = ws_formulas.cell(12, 20).value or ""  # T12
    dc_formula = ws_formulas.cell(15, 20).value or ""  # T15
    for m in re.findall(r'K(\d+)', cc_formula):
        cc_rows.add(int(m))
    for m in re.findall(r'K(\d+)', dc_formula):
        dc_rows.add(int(m))

    expenses = []
    current_date = None
    current_date_label = ""
    day_counter = 0

    # Track summary data from columns S-W
    summary = {"total_all": {}, "credit_card": {}, "debit_card": {}}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        date_cell = row[1]  # Column B
        transaction_cell = row[2]  # Column C
        details_cell = row[3]  # Column D
        total_cell = row[4]  # Column E

        jes_cell = row[10]  # Column K
        cha_cell = row[11]  # Column L
        joyce_cell = row[12]  # Column M
        joh_cell = row[13]  # Column N

        # Update current date if present
        # Dates in Excel are stored as mmm-yy format where the year's last 2 digits = day of May
        # e.g., 2016-05-01 displays as "May-16" meaning May 16th
        if date_cell.value is not None:
            if isinstance(date_cell.value, datetime):
                day_counter += 1
                # The year (e.g. 2016) last 2 digits = actual day of month (16)
                actual_day = date_cell.value.year % 100
                current_date_label = f"Day {day_counter} (May {actual_day})"
            elif isinstance(date_cell.value, str) and date_cell.value.strip().upper() == "OTHERS":
                current_date_label = "Others (Pre-booked)"
                day_counter += 1

        # Skip rows with no transaction info
        transaction = transaction_cell.value
        if transaction is None:
            # Check if it's a detail/sub-item row (has details or amounts)
            if details_cell.value is not None or total_cell.value is not None:
                continue
            continue

        transaction = str(transaction).strip()
        if not transaction:
            continue

        # Skip division rows like "/2", "/3", "/4"
        if transaction.startswith("/") and transaction[1:].strip().isdigit():
            continue

        # Get individual amounts
        jes_amount = jes_cell.value if isinstance(jes_cell.value, (int, float)) else 0
        cha_amount = cha_cell.value if isinstance(cha_cell.value, (int, float)) else 0
        joyce_amount = joyce_cell.value if isinstance(joyce_cell.value, (int, float)) else 0
        joh_amount = joh_cell.value if isinstance(joh_cell.value, (int, float)) else 0

        # Only include rows that have actual per-person amounts (K/L/M/N)
        # Rows with only column E values are sub-items/descriptions - skip them
        if jes_amount == 0 and cha_amount == 0 and joyce_amount == 0 and joh_amount == 0:
            continue

        # Calculate total from individual amounts (more reliable than column E)
        total_amount = jes_amount + cha_amount + joyce_amount + joh_amount

        # Determine mode of payment from CC/DC formula references
        excel_row = row[0].row
        if excel_row in cc_rows:
            payment = "CC"
        elif excel_row in dc_rows:
            payment = "DC"
        else:
            payment = "Others"

        expenses.append({
            "day": current_date_label,
            "dayNum": day_counter,
            "transaction": transaction,
            "total": round(total_amount, 2),
            "jes": round(jes_amount, 2),
            "cha": round(cha_amount, 2),
            "joyce": round(joyce_amount, 2),
            "joh": round(joh_amount, 2),
            "payment": payment,
        })

    # Parse summary data from columns S-W (rows 8-15)
    summary = {
        "total_all": {"jes": 0, "cha": 0, "joyce": 0, "joh": 0},
        "credit_card": {"total": 0, "jes": 0, "cha": 0, "joyce": 0, "joh": 0},
        "debit_card": {"total": 0, "jes": 0, "cha": 0, "joyce": 0, "joh": 0},
    }

    # Row 9: Total of all expenses (cols T, U = JJ, CHA)
    row9 = [ws.cell(row=9, column=c).value for c in range(20, 24)]
    summary["total_all"]["jes"] = round(row9[0] or 0, 2)
    summary["total_all"]["cha"] = round(row9[1] or 0, 2)
    summary["total_all"]["joyce"] = round(row9[2] or 0, 2)
    summary["total_all"]["joh"] = round(row9[3] or 0, 2)

    # Row 12: Credit card expenses (cols S-W)
    row12 = [ws.cell(row=12, column=c).value for c in range(19, 24)]
    summary["credit_card"]["total"] = round(row12[0] or 0, 2)
    summary["credit_card"]["jes"] = round(row12[1] or 0, 2)
    summary["credit_card"]["cha"] = round(row12[2] or 0, 2)
    summary["credit_card"]["joyce"] = round(row12[3] or 0, 2)
    summary["credit_card"]["joh"] = round(row12[4] or 0, 2)

    # Row 15: Debit card withdrawals (cols S-W)
    row15 = [ws.cell(row=15, column=c).value for c in range(19, 24)]
    summary["debit_card"]["total"] = round(row15[0] or 0, 2)
    summary["debit_card"]["jes"] = round(row15[1] or 0, 2)
    summary["debit_card"]["cha"] = round(row15[2] or 0, 2)
    summary["debit_card"]["joyce"] = round(row15[3] or 0, 2)
    summary["debit_card"]["joh"] = round(row15[4] or 0, 2)

    wb.close()
    return {"expenses": expenses, "summary": summary}


class ReportHandler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/" or path == "/index.html":
            try:
                self.send_response(200)
                self.send_header("Content-type", "text/html; charset=utf-8")
                self.end_headers()
                with open(HTML_FILE, "r", encoding="utf-8") as f:
                    self.wfile.write(f.read().encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-type", "text/plain")
                self.end_headers()
                self.wfile.write(f"Error: {e}".encode("utf-8"))
        elif path == "/api/data":
            try:
                expenses = parse_excel()
                self.send_response(200)
                self.send_header("Content-type", "application/json")
                self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
                self.end_headers()
                self.wfile.write(json.dumps(expenses).encode("utf-8"))
            except Exception as e:
                self.send_response(500)
                self.send_header("Content-type", "application/json")
                self.end_headers()
                self.wfile.write(json.dumps({"error": str(e)}).encode("utf-8"))
        else:
            super().do_GET()

    def log_message(self, format, *args):
        print(f"[{datetime.now().strftime('%H:%M:%S')}] {args[0]}")


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    server = http.server.HTTPServer(("0.0.0.0", PORT), ReportHandler)
    print(f"Japan Trip Expense Report Server")
    print(f"Local access:   http://localhost:{PORT}")
    print(f"Network access: http://{local_ip}:{PORT}")
    print(f"Reading from: {EXCEL_FILE}")
    print(f"Press Ctrl+C to stop")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped.")
        server.server_close()
