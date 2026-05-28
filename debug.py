import openpyxl
from datetime import datetime

wb = openpyxl.load_workbook('temp_copy.xlsx', data_only=True)
ws = wb['Japan Trip']

expenses = []
current_date_label = ''
day_counter = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    date_cell = row[1]
    transaction_cell = row[2]
    total_cell = row[4]
    jes_cell = row[10]
    cha_cell = row[11]
    joyce_cell = row[12]
    joh_cell = row[13]

    if date_cell.value is not None:
        if isinstance(date_cell.value, datetime):
            day_counter += 1
            actual_day = date_cell.value.year % 100
            current_date_label = f'Day {day_counter} (May {actual_day})'
        elif isinstance(date_cell.value, str) and date_cell.value.strip().upper() == 'OTHERS':
            current_date_label = 'Others (Pre-booked)'
            day_counter += 1

    transaction = transaction_cell.value
    if transaction is None:
        continue
    transaction = str(transaction).strip()
    if not transaction:
        continue
    if transaction.startswith('/') and transaction[1:].strip().isdigit():
        continue

    jes = jes_cell.value if isinstance(jes_cell.value, (int, float)) else 0
    cha = cha_cell.value if isinstance(cha_cell.value, (int, float)) else 0
    joyce = joyce_cell.value if isinstance(joyce_cell.value, (int, float)) else 0
    joh = joh_cell.value if isinstance(joh_cell.value, (int, float)) else 0

    # Only include rows that have actual per-person amounts (K/L/M/N)
    if jes == 0 and cha == 0 and joyce == 0 and joh == 0:
        continue

    # Check if name is a person name - look up proper store name
    person_names = ["jes", "cha", "joyce", "joh"]
    lower_txn = transaction.lower()
    is_person = any(lower_txn == p or lower_txn.startswith(p + " -") or lower_txn.startswith(p + " ") for p in person_names)
    display_name = transaction
    if is_person:
        excel_row = row[0].row
        for back_r in range(excel_row - 1, 1, -1):
            back_txn = ws.cell(back_r, 3).value
            if back_txn is None:
                continue
            back_str = str(back_txn).strip()
            if not back_str or back_str.startswith("/"):
                continue
            back_lower = back_str.lower()
            if any(back_lower == p or back_lower.startswith(p + " -") or back_lower.startswith(p + " ") for p in person_names):
                continue
            bk = ws.cell(back_r, 11).value or 0
            bl = ws.cell(back_r, 12).value or 0
            bm = ws.cell(back_r, 13).value or 0
            bn = ws.cell(back_r, 14).value or 0
            back_sum = (bk if isinstance(bk, (int, float)) else 0) + (bl if isinstance(bl, (int, float)) else 0) + (bm if isinstance(bm, (int, float)) else 0) + (bn if isinstance(bn, (int, float)) else 0)
            if back_sum == 0:
                display_name = back_str
            break

    total = jes + cha + joyce + joh

    expenses.append({'day': current_date_label, 'txn': display_name, 'total': total, 'jes': jes, 'cha': cha, 'joyce': joyce, 'joh': joh})

print(f'Total transactions: {len(expenses)}')
grand_total = sum(e['total'] for e in expenses)
jes_total = sum(e['jes'] for e in expenses)
cha_total = sum(e['cha'] for e in expenses)

print(f'Grand Total (sum of all): {grand_total:.2f}')
print(f'JES total: {jes_total:.2f}')
print(f'CHA total: {cha_total:.2f}')
print(f'JOYCE total: {sum(e["joyce"] for e in expenses):.2f}')
print(f'JOH total: {sum(e["joh"] for e in expenses):.2f}')
print()
print('Excel summary says:')
print(f'  TOTAL ALL - JJ: {ws.cell(9,20).value}, CHA: {ws.cell(9,21).value}')
print(f'  CREDIT CARD total: {ws.cell(12,19).value}')
print(f'  DEBIT CARD total: {ws.cell(15,19).value}')
print()
print('--- All parsed transactions ---')
for e in expenses:
    print(f"  {e['day']:25s} | {e['txn']:40s} | Total: {e['total']:>10.2f} | JES: {e['jes']:>9.2f} | CHA: {e['cha']:>9.2f}")
